import json
import re

import openpyxl
from openpyxl.cell import MergedCell, Cell
from openpyxl.workbook.child import _WorkbookChild as Sheet

FI = "ФІ"
FEN = "ФЕН"
LECTURE = "лекція"
FIRST_DISCIPLINE_OFFSET = 11

output = {
    "ФІ": {},
    "ФЕН": {}
}


def get_cell_value(cell: Cell | MergedCell) -> str:
    return str(cell.value).replace('\n', '') if cell.value else None


def read_merged_cell(workbook_data, row: int, column: int) -> str:
    cell = workbook_data.cell(row=row, column=column)
    if isinstance(cell, MergedCell):
        for merged_range in workbook_data.merged_cells.ranges:
            if cell.coordinate in merged_range:

                cell = workbook_data.cell(row=merged_range.min_row,
                                          column=merged_range.min_col)
                break
    return get_cell_value(cell)


def get_specialties(raw_specialties: str) -> list[str]:
    if '"' in raw_specialties:
        specialties_array = raw_specialties.split('"')

        # We split specialties by quotation mark (") so every second element in that array is inside quotes
        return specialties_array[1::2]
    else:
        raw_specialties_array = raw_specialties.split('«')
        specialties_array = []

        for specialty in raw_specialties_array:
            if '»' in specialty:
                specialties_array.append(specialty.split('»')[0])

        return specialties_array


def get_abbreviated_specialties(discipline_name: str) -> list[str] | None:
    abbrreviated_specialties = []
    for sub_string in discipline_name.split('('):
        if ')' in sub_string:
            specialties = sub_string.split(')')[0]
            clean_specialties = re.sub(r'[^\w+,]+', '', specialties)
            if '+' in specialties:
                abbrreviated_specialties.extend(clean_specialties.split('+'))
            else:
                abbrreviated_specialties.extend(clean_specialties.split(','))

    return abbrreviated_specialties


def create_schedule_structure(sheet: Sheet):
    specialties = get_specialties(get_cell_value(sheet.cell(7, 1)))

    structure = {}

    for row in range(FIRST_DISCIPLINE_OFFSET, sheet.max_row):
        for specialty in specialties:
            if not specialty in structure:
                structure[specialty] = {}

            discipline_name = get_cell_value(sheet.cell(row, 3))

            if not discipline_name:
                continue

            group = get_cell_value(sheet.cell(row, 4))
            if LECTURE in group.lower():
                group = LECTURE
            else:
                group = re.sub('\D', '', group)

            def add_discipline():
                if not discipline_name in structure[specialty]:
                    structure[specialty][discipline_name] = {}

                structure[specialty][discipline_name][group] = {
                    "час": read_merged_cell(sheet, row, 2),
                    "тижні": get_cell_value(sheet.cell(row, 5)),
                    "аудиторія": get_cell_value(sheet.cell(row, 6)),
                    "день тижня": read_merged_cell(sheet, row, 1)
                }

            abbreviated_specialties = get_abbreviated_specialties(
                discipline_name)

            all_specialties = False
            if len(abbreviated_specialties) == 0:
                all_specialties = True

            if all_specialties:
                add_discipline()

            for abbreviated_specialty in abbreviated_specialties:
                if abbreviated_specialty.lower() in specialty.lower():
                    add_discipline()

    return structure


def main():
    workbook_fen = openpyxl.load_workbook("FEN.xlsx")
    current_sheet = workbook_fen.active
    output[FEN] = create_schedule_structure(current_sheet)

    workbook_ipz = openpyxl.load_workbook("IPZ.xlsx")
    current_sheet = workbook_ipz.active
    output[FI] = create_schedule_structure(current_sheet)

    with open('output.json', 'w', encoding='utf-8') as f:
        json.dump(output, f, ensure_ascii=False, indent=4)


if __name__ == '__main__':
    main()
